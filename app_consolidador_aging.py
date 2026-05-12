"""
🌎 CONSOLIDADOR DE AGING - CON EXCLUSIONES AMPLIADAS
Aplicación web para consolidar archivos aging excluyendo PERSONA/CONSUMIDOR y clientes específicos
"""

import streamlit as st
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import numbers
import pandas as pd
import re
from datetime import datetime, date
from io import BytesIO
import tempfile
import os

# Configuración de la página
st.set_page_config(
    page_title="Consolidador Aging",
    page_icon="🌎",
    layout="wide"
)

# ====================================================================
# LISTA DE CLIENTES A EXCLUIR
# ====================================================================

CLIENTES_EXCLUIR = {
    # Argentina
    'Consumidor Final (AR)',
    'Responsable Inscripto (AR)', 
    'Monotributista (AR)',
    'CONSUMIDOR FINAL COUPON (AR)',
    'RESPONSABLE INCRIPTO B COUPON (AR)',
    'MONTRIBUTO COUPON (AR)',
    'EXENTO COUPON (AR)',
    'RESPONSABLE INSCRIPTO COUPON (AR)',
    'Exento (AR)',
    'CONSUMIDOR FINAL COUPON USD (AR)',
    
    # Brasil - Existentes
    'Pessoa Fisica (BR)',
    'Pessoa Juridica (BR)', 
    'Estrangeiros (BR)',
    'Nao Informado (BR)',
    'PESSOA FISICA COUPON (BR)',
    'PESSOA JURIDICA COUPON (BR)',
    'PESSOA EXT COUPON (BR)',
    
    # Brasil - Nuevos (más genéricos)
    'Pessoa',
    'Estrangeiros', 
    'Nao Informado',
    
    # Multipos existentes
    'MULTIPOS AR (BR)',
    'MULTIPOS MX (BR)', 
    'MULTIPOS CL (BR)',
}

PALABRAS_EXCLUIR = {
    'PERSONA',
    'CONSUMIDOR',
    'MULTIPOS'
}

# ====================================================================
# FUNCIONES DE PROCESAMIENTO
# ====================================================================

def extraer_entidad(nombre_archivo):
    """Extrae el nombre del país del nombre del archivo"""
    nombre_base = os.path.splitext(nombre_archivo)[0]
    
    nombre_limpio = nombre_base
    palabras_eliminar = ['aging', 'filtrado', 'sin_persona', 'consolidado']
    
    for palabra in palabras_eliminar:
        nombre_limpio = re.sub(rf'{palabra}[_\s-]*', '', nombre_limpio, flags=re.IGNORECASE)
        nombre_limpio = re.sub(rf'[_\s-]*{palabra}', '', nombre_limpio, flags=re.IGNORECASE)
    
    nombre_limpio = re.sub(r'[_\s-]*\d+$', '', nombre_limpio)
    nombre_limpio = nombre_limpio.strip('_- ')
    
    if not nombre_limpio:
        partes = nombre_base.split('_')
        for parte in partes:
            if parte.lower() not in palabras_eliminar and len(parte) > 0:
                nombre_limpio = parte
                break
    
    return nombre_limpio.upper() if nombre_limpio else 'DESCONOCIDO'

def debe_excluir_registro(row):
    """
    Determina si un registro debe ser excluido
    Retorna: (debe_excluir, razon)
    """
    
    # Convertir fila a texto para buscar palabras clave
    row_text = ' '.join([str(cell).upper() if cell is not None else '' for cell in row])
    
    # 1. Verificar palabras clave (PERSONA, CONSUMIDOR, MULTIPOS)
    for palabra in PALABRAS_EXCLUIR:
        if palabra in row_text:
            return True, f"Contiene palabra: {palabra}"
    
    # 2. Verificar nombres exactos de clientes (buscar en toda la fila)
    for cliente_exacto in CLIENTES_EXCLUIR:
        if cliente_exacto.upper() in row_text:
            return True, f"Cliente exacto: {cliente_exacto}"
    
    # 3. Si no se encuentra nada, conservar el registro
    return False, "Registro válido"

def es_fila_valida(row, min_columnas=5):
    """Valida que la fila tenga datos reales"""
    if not row:
        return False
    
    primeras_columnas = row[:10]
    datos_importantes = sum(1 for cell in primeras_columnas 
                           if cell is not None and str(cell).strip() not in ['', 'None', 'nan'])
    
    if datos_importantes < min_columnas:
        return False
    
    primera_col = str(row[0]).strip().upper() if row[0] is not None else ''
    
    palabras_excluir = [
        'PROGRAMA', 'FECHA Y', 'REPORTE', 'TOTAL', 'SUBTOTAL',
        'RESUMEN', 'NOTA:', 'OBSERVACION', 'CONCURRENTE'
    ]
    
    if any(palabra in primera_col for palabra in palabras_excluir):
        return False
    
    if primera_col in ['COLOMBIA', 'MEXICO', 'BRASIL', 'ARGENTINA', 'CHILE', 'PERU']:
        if datos_importantes <= 1:
            return False
    
    return True

def es_fecha(valor):
    """Detecta si un valor es una fecha"""
    if valor is None:
        return False
    
    if isinstance(valor, (date, datetime)):
        return True
    
    if isinstance(valor, str):
        patrones_fecha = [
            r'\d{4}-\d{2}-\d{2}',
            r'\d{2}/\d{2}/\d{4}',
            r'\d{2}-\d{2}-\d{4}',
        ]
        for patron in patrones_fecha:
            if re.match(patron, str(valor)):
                return True
    
    return False

def detectar_columnas_fecha(headers, filas_muestra):
    """Detecta qué columnas contienen fechas"""
    columnas_fecha = set()
    
    palabras_fecha = ['fecha', 'date', 'emissao', 'vencimento', 'vto', 'emision', 'vencimiento']
    for idx, header in enumerate(headers):
        if header and any(palabra in str(header).lower() for palabra in palabras_fecha):
            columnas_fecha.add(idx)
    
    for fila in filas_muestra[:50]:
        for idx, valor in enumerate(fila):
            if es_fecha(valor):
                columnas_fecha.add(idx)
    
    return columnas_fecha

def limpiar_nombres_columnas(headers):
    """Limpia y asegura que no haya nombres de columnas duplicados"""
    headers_limpios = []
    contadores = {}
    
    for header in headers:
        # Convertir None a string
        if header is None:
            header = "Columna_Sin_Nombre"
        else:
            header = str(header).strip()
        
        # Si está vacío, usar nombre genérico
        if not header:
            header = "Columna_Vacia"
        
        # Si ya existe, agregar número
        header_original = header
        if header in contadores:
            contadores[header] += 1
            header = f"{header_original}_{contadores[header]}"
        else:
            contadores[header] = 0
        
        headers_limpios.append(header)
    
    return headers_limpios

def procesar_archivo(archivo_bytes, nombre_archivo, progress_bar, status_text):
    """Procesa un archivo aging con exclusiones ampliadas"""
    
    entidad = extraer_entidad(nombre_archivo)
    status_text.text(f"📂 Procesando: {nombre_archivo} → Entidad: {entidad}")
    
    # Guardar temporalmente el archivo
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
        tmp_file.write(archivo_bytes)
        tmp_path = tmp_file.name
    
    try:
        # Abrir archivo
        wb_read = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
        ws = wb_read.active
        
        # Leer headers y limpiarlos
        headers_originales = [cell.value for cell in ws[2]]
        headers_limpios = limpiar_nombres_columnas(headers_originales)
        
        # VERIFICAR si ya existe una columna "Entidad"
        if "Entidad" in headers_limpios:
            # Si ya existe, no agregar otra
            headers_nuevos = headers_limpios
            entidad_ya_existe = True
        else:
            # Si no existe, agregar al inicio
            headers_nuevos = ['Entidad'] + headers_limpios
            entidad_ya_existe = False
        
        # Procesar filas
        filas_filtradas = []
        total_procesadas = 0
        excluidos_por_palabra = 0
        excluidos_por_cliente = 0
        excluidos_invalidos = 0
        conservados = 0
        filas_consecutivas_vacias = 0
        
        # Contadores por razón de exclusión
        razones_exclusion = {}
        
        max_rows = ws.max_row
        
        for idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            total_procesadas += 1
            
            # Actualizar barra de progreso
            if total_procesadas % 500 == 0:
                progress = min((idx - 2) / max_rows, 1.0)
                progress_bar.progress(progress)
            
            # Validar fila
            if not es_fila_valida(row, min_columnas=5):
                excluidos_invalidos += 1
                filas_consecutivas_vacias += 1
                
                if filas_consecutivas_vacias > 100:
                    break
                
                continue
            
            filas_consecutivas_vacias = 0
            
            # Verificar si debe excluirse
            debe_excluir, razon = debe_excluir_registro(row)
            
            if debe_excluir:
                # Contar razones de exclusión
                if razon in razones_exclusion:
                    razones_exclusion[razon] += 1
                else:
                    razones_exclusion[razon] = 1
                
                if "palabra" in razon.lower():
                    excluidos_por_palabra += 1
                else:
                    excluidos_por_cliente += 1
                continue
            
            # Agregar entidad solo si no existe ya
            if entidad_ya_existe:
                fila_con_entidad = list(row)
            else:
                fila_con_entidad = [entidad] + list(row)
            
            filas_filtradas.append(fila_con_entidad)
            conservados += 1
        
        wb_read.close()
        
        # Detectar columnas con fechas
        columnas_fecha = detectar_columnas_fecha(headers_nuevos, filas_filtradas)
        
        return {
            'entidad': entidad,
            'headers': headers_nuevos,
            'filas': filas_filtradas,
            'columnas_fecha': columnas_fecha,
            'stats': {
                'procesadas': total_procesadas,
                'conservadas': conservados,
                'excluidas_palabra': excluidos_por_palabra,
                'excluidas_cliente': excluidos_por_cliente,
                'excluidas_invalidas': excluidos_invalidos,
                'razones_exclusion': razones_exclusion
            }
        }
    
    finally:
        # Limpiar archivo temporal
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)

def columna_letra(idx):
    """Convierte índice numérico a letra de columna Excel"""
    letra = ""
    idx += 1
    while idx > 0:
        idx -= 1
        letra = chr(65 + (idx % 26)) + letra
        idx //= 26
    return letra

def crear_excel_consolidado(resultados):
    """Crea el archivo Excel consolidado"""
    
    wb_salida = Workbook()
    ws_salida = wb_salida.active
    ws_salida.title = "Consolidado"
    
    headers_escritos = False
    fila_actual = 1
    columnas_fecha_global = set()
    
    # Escribir datos
    for resultado in resultados:
        # Headers (solo una vez)
        if not headers_escritos:
            ws_salida.append(resultado['headers'])
            fila_actual += 1
            headers_escritos = True
        
        # Actualizar columnas fecha
        columnas_fecha_global.update(resultado['columnas_fecha'])
        
        # Escribir filas
        for fila in resultado['filas']:
            ws_salida.append(fila)
            fila_actual += 1
    
    # Aplicar formato de fecha
    if columnas_fecha_global:
        for col_idx in columnas_fecha_global:
            col_letra = columna_letra(col_idx)
            
            for fila_num in range(2, fila_actual + 1):
                celda = ws_salida[f'{col_letra}{fila_num}']
                celda.number_format = 'DD/MM/YYYY'
    
    # Guardar en BytesIO
    output = BytesIO()
    wb_salida.save(output)
    output.seek(0)
    
    return output, columnas_fecha_global

# ====================================================================
# INTERFAZ STREAMLIT
# ====================================================================

st.title("🌎 Consolidador de Aging")
st.markdown("### Excluye PERSONA/CONSUMIDOR y clientes específicos de AR/BR")

st.divider()

# Instrucciones mejoradas
with st.expander("📋 Instrucciones de uso", expanded=False):
    st.markdown("""
    **Cómo usar esta aplicación:**
    
    1. 📤 **Sube tus archivos aging** (puedes subir varios a la vez)
    2. ⚙️ **El sistema excluye automáticamente:**
       - Registros con palabras: PERSONA, CONSUMIDOR, MULTIPOS
       - Clientes específicos de Argentina y Brasil (ver lista abajo)
       - Valida que las filas tengan datos reales
       - Agrega columna "Entidad" al inicio (si no existe)
       - Aplica formato de fecha corta (dd/mm/yyyy)
    3. 📥 **Descarga el archivo consolidado**
    
    **🚫 Clientes excluidos automáticamente:**
    
    **Argentina (AR):**
    - Consumidor Final, Responsable Inscripto, Monotributista
    - CONSUMIDOR FINAL COUPON, RESPONSABLE INSCRIPTO COUPON
    - EXENTO COUPON, Exento, etc.
    
    **Brasil (BR):**
    - Pessoa Fisica, Pessoa Juridica, Estrangeiros
    - Nao Informado, MULTIPOS, etc.
    
    **Características:**
    - ✅ Consolidación multi-país automática
    - ✅ Formato de fechas automático
    - ✅ Estadísticas detalladas por razón de exclusión
    - ✅ Validación de filas
    """)

# Mostrar clientes que se excluyen
with st.expander("🚫 Lista completa de clientes excluidos", expanded=False):
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("**Argentina (AR):**")
        ar_clients = [c for c in CLIENTES_EXCLUIR if '(AR)' in c]
        for client in sorted(ar_clients):
            st.write(f"• {client}")
    
    with col2:
        st.markdown("**Brasil (BR) y otros:**")
        br_clients = [c for c in CLIENTES_EXCLUIR if '(BR)' in c or c in ['Pessoa', 'Estrangeiros', 'Nao Informado']]
        for client in sorted(br_clients):
            st.write(f"• {client}")
    
    st.markdown("**Palabras clave excluidas en cualquier parte:**")
    for palabra in sorted(PALABRAS_EXCLUIR):
        st.write(f"• {palabra}")

# Subir archivos
st.subheader("📤 1. Subir archivos Aging")

uploaded_files = st.file_uploader(
    "Selecciona uno o más archivos Excel (.xlsx)",
    type=['xlsx'],
    accept_multiple_files=True,
    help="Puedes seleccionar múltiples archivos para consolidar"
)

if uploaded_files:
    st.success(f"✅ {len(uploaded_files)} archivo(s) seleccionado(s)")
    
    # Mostrar archivos
    with st.expander("📂 Archivos cargados", expanded=True):
        for file in uploaded_files:
            entidad = extraer_entidad(file.name)
            col1, col2, col3 = st.columns([3, 2, 1])
            col1.write(f"📄 {file.name}")
            col2.write(f"🌎 Entidad: **{entidad}**")
            col3.write(f"{file.size / 1024 / 1024:.2f} MB")
    
    st.divider()
    
    # Botón procesar
    st.subheader("⚙️ 2. Procesar archivos")
    
    if st.button("🚀 Consolidar archivos", type="primary", use_container_width=True):
        
        # Contenedores para progreso
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        resultados = []
        stats_totales = {
            'conservadas': 0,
            'excluidas_palabra': 0,
            'excluidas_cliente': 0,
            'excluidas_invalidas': 0
        }
        
        # Procesar cada archivo
        for i, uploaded_file in enumerate(uploaded_files):
            status_text.text(f"Procesando archivo {i+1} de {len(uploaded_files)}...")
            
            archivo_bytes = uploaded_file.read()
            resultado = procesar_archivo(
                archivo_bytes, 
                uploaded_file.name,
                progress_bar,
                status_text
            )
            
            resultados.append(resultado)
            
            # Acumular estadísticas
            stats_totales['conservadas'] += resultado['stats']['conservadas']
            stats_totales['excluidas_palabra'] += resultado['stats']['excluidas_palabra']
            stats_totales['excluidas_cliente'] += resultado['stats']['excluidas_cliente']
            stats_totales['excluidas_invalidas'] += resultado['stats']['excluidas_invalidas']
        
        progress_bar.progress(1.0)
        status_text.text("✅ Procesamiento completado")
        
        st.divider()
        
        # Mostrar estadísticas
        st.subheader("📊 3. Resultados")
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric(
                "✅ Registros conservados",
                f"{stats_totales['conservadas']:,}"
            )
        
        with col2:
            st.metric(
                "🚫 Excluidos (Palabras)",
                f"{stats_totales['excluidas_palabra']:,}",
                help="PERSONA, CONSUMIDOR, MULTIPOS"
            )
        
        with col3:
            st.metric(
                "🚫 Excluidos (Clientes)",
                f"{stats_totales['excluidas_cliente']:,}",
                help="Clientes específicos AR/BR"
            )
        
        with col4:
            st.metric(
                "🗑️ Excluidos (Inválidos)",
                f"{stats_totales['excluidas_invalidas']:,}",
                help="Filas vacías o con pocos datos"
            )
        
        # Estadísticas por país
        st.markdown("#### 📈 Registros por país:")
        
        df_stats = pd.DataFrame([
            {
                'País': r['entidad'],
                'Conservados': f"{r['stats']['conservadas']:,}",
                'Excluidos Palabras': f"{r['stats']['excluidas_palabra']:,}",
                'Excluidos Clientes': f"{r['stats']['excluidas_cliente']:,}",
                'Excluidos Inválidos': f"{r['stats']['excluidas_invalidas']:,}"
            }
            for r in resultados
        ])
        
        st.dataframe(df_stats, use_container_width=True, hide_index=True)
        
        # Mostrar razones detalladas de exclusión
        with st.expander("📋 Detalle de exclusiones por razón", expanded=False):
            for resultado in resultados:
                st.markdown(f"**{resultado['entidad']}:**")
                razones = resultado['stats']['razones_exclusion']
                for razon, cantidad in sorted(razones.items(), key=lambda x: x[1], reverse=True):
                    st.write(f"• {razon}: {cantidad:,} registros")
                st.divider()
        
        st.divider()
        
        # Crear archivo consolidado
        st.subheader("💾 4. Descargar resultado")
        
        with st.spinner("Generando archivo consolidado..."):
            excel_consolidado, columnas_fecha = crear_excel_consolidado(resultados)
        
        # Información adicional
        col1, col2 = st.columns(2)
        
        with col1:
            st.info(f"📋 Total de filas: **{stats_totales['conservadas']:,}**")
        
        with col2:
            st.info(f"📅 Columnas con fecha formateada: **{len(columnas_fecha)}**")
        
        # Botón de descarga
        nombre_salida = "Aging_CONSOLIDADO.xlsx" if len(uploaded_files) > 1 else f"Aging_{resultados[0]['entidad']}_FILTRADO.xlsx"
        
        st.download_button(
            label="📥 Descargar archivo consolidado",
            data=excel_consolidado,
            file_name=nombre_salida,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True
        )
        
        st.success("✅ ¡Consolidación completada exitosamente!")
        
        # Vista previa SIN ERRORES
        with st.expander("👁️ Vista previa de datos", expanded=False):
            if resultados:
                try:
                    # Crear dataframe para preview de manera segura
                    preview_data = []
                    headers_para_preview = resultados[0]['headers']
                    
                    for resultado in resultados[:3]:  # Máximo 3 países
                        for fila in resultado['filas'][:10]:  # Máximo 10 filas por país
                            # Asegurar que la fila tenga la cantidad correcta de columnas
                            fila_ajustada = list(fila)
                            while len(fila_ajustada) < len(headers_para_preview):
                                fila_ajustada.append("")
                            preview_data.append(fila_ajustada[:len(headers_para_preview)])
                    
                    if preview_data:
                        # Crear DataFrame de manera segura
                        df_preview = pd.DataFrame(preview_data, columns=headers_para_preview)
                        
                        # Verificar que no hay columnas duplicadas
                        if df_preview.columns.duplicated().any():
                            st.warning("⚠️ Se detectaron columnas duplicadas. Mostrando solo las primeras 5 columnas:")
                            st.dataframe(df_preview.iloc[:, :5].head(20), use_container_width=True)
                        else:
                            st.dataframe(df_preview.head(50), use_container_width=True)
                        
                        st.caption("Mostrando muestra del resultado consolidado")
                        
                except Exception as e:
                    st.warning(f"⚠️ No se pudo mostrar la vista previa, pero el archivo se descargó correctamente.")

else:
    st.info("👆 Por favor, sube uno o más archivos para comenzar")

# Footer
st.divider()
st.markdown("""
<div style='text-align: center; color: gray; font-size: 0.9em;'>
    🌎 Consolidador de Aging v1.2 - CON EXCLUSIONES AMPLIADAS | Desarrollado con Streamlit
</div>
""", unsafe_allow_html=True)
